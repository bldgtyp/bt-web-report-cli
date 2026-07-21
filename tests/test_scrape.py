import csv
import json
import os
from dataclasses import dataclass
from pathlib import Path

import pytest
import yaml
from click.testing import CliRunner
from openpyxl import Workbook

from bt_web_report_cli.__main__ import main
from bt_web_report_cli.io.workbook_openpyxl import OpenpyxlWorkbookReader
from bt_web_report_cli.phpp.write import REPORT_CSV_TABLES
from bt_web_report_schemas.phpp.models import ClimateMonthlySchema, RoomVentilationSchema, WorkbookSchema

FIXTURE_DIR = Path(os.environ.get("BTWR_TEST_PHPP_DIR", Path(__file__).resolve().parents[2] / "test-files" / "phpp"))
VANDAM_DIR = FIXTURE_DIR / "2606-Vandam-St"
LINDE_DIR = FIXTURE_DIR / "2524-Linde-Residence"
VANDAM_FIXTURE = VANDAM_DIR / "2606-29-Vandam-St.xlsx"
LINDE_FIXTURE = LINDE_DIR / "2524-Linde-Residence-250709.xlsx"


@dataclass(frozen=True)
class ScrapeOutput:
    data_dir: Path
    command_output: str


@pytest.fixture(scope="session")
def scraped_vandam(tmp_path_factory: pytest.TempPathFactory) -> ScrapeOutput:
    return _scrape_fixture(VANDAM_FIXTURE, tmp_path_factory.mktemp("vandam-scrape") / "data")


@pytest.fixture(scope="session")
def scraped_linde(tmp_path_factory: pytest.TempPathFactory) -> ScrapeOutput:
    return _scrape_fixture(LINDE_FIXTURE, tmp_path_factory.mktemp("linde-scrape") / "data")


def test_vandam_golden_csvs(scraped_vandam: ScrapeOutput) -> None:
    _assert_golden_csvs(scraped_vandam.data_dir, VANDAM_DIR / "scrape-output")


def test_linde_golden_csvs(scraped_linde: ScrapeOutput) -> None:
    _assert_golden_csvs(scraped_linde.data_dir, LINDE_DIR / "scrape-output")


def test_scrape_fixture_writes_manifest_and_variants_csv(scraped_vandam: ScrapeOutput) -> None:
    output_dir = scraped_vandam.data_dir

    assert "scraped PHPP 10.6: 5 variants, recommended=enerphit_by_demand" in scraped_vandam.command_output

    manifest = json.loads((output_dir / "manifest.json").read_text())
    assert manifest["phpp_version"] == "10.6"
    assert manifest["recommended_variant_id"] == "enerphit_by_demand"
    assert [variant["id"] for variant in manifest["variants"]] == [
        "code_minimum",
        "improved_envelope",
        "improved_hvac",
        "enerphit_by_component",
        "enerphit_by_demand",
    ]
    assert manifest["variants"][-1]["recommended"] is True
    assert manifest["source_workbook"]["size_bytes"] > 1_000_000

    rows = list(csv.DictReader((output_dir / "variants.csv").open()))
    tfa = [row for row in rows if row["field_id"] == "geometry.tfa"]
    assert len(tfa) == 5
    assert tfa[0]["variant_id"] == "code_minimum"
    assert tfa[0]["units"] == "m2"
    assert float(tfa[0]["value"]) == pytest.approx(290.2644471258237)

    climate_rows = list(csv.DictReader((output_dir / "climate-monthly.csv").open()))
    assert len(climate_rows) == 96
    jan_exterior = [row for row in climate_rows if row["month"] == "jan" and row["metric"] == "exterior_temperature"][0]
    assert jan_exterior["units"] == "degF"
    assert float(jan_exterior["value"]) == 32.54
    jan_north = [row for row in climate_rows if row["month"] == "jan" and row["orientation"] == "north"][0]
    assert jan_north["units"] == "kWh/ft2"
    assert round(float(jan_north["value"]), 6) == round(16 / 10.76391042, 6)

    airflow_rows = list(csv.DictReader((output_dir / "room-airflows.csv").open()))
    assert airflow_rows[-1]["row_type"] == "total"
    assert airflow_rows[-1]["room_name"] == "Totals"
    assert any(row["room_name"] == "104-KITCHEN" for row in airflow_rows)

    building_metric_rows = list(csv.DictReader((output_dir / "building-metrics.csv").open()))
    tfa_ft2 = [
        row
        for row in building_metric_rows
        if row["metric"] == "treated_floor_area" and row["variant_id"] == "code_minimum"
    ][0]
    assert tfa_ft2["units"] == "ft2"
    assert round(float(tfa_ft2["value"]), 6) == round(290.2644471258237 * 10.76391042, 6)

    certification_rows = list(csv.DictReader((output_dir / "certification.csv").open()))
    heat_demand = [
        row
        for row in certification_rows
        if row["metric"] == "heat_demand" and row["role"] == "result" and row["variant_id"] == "code_minimum"
    ][0]
    assert heat_demand["units"] == "kWh"
    assert round(float(heat_demand["value"]), 6) == round(63.764478348563266 * 290.2644471258237, 6)
    heat_demand_limit = [
        row
        for row in certification_rows
        if row["metric"] == "heat_demand" and row["role"] == "limit" and row["variant_id"] == "enerphit_by_demand"
    ][0]
    assert heat_demand_limit["units"] == "kWh"
    assert round(float(heat_demand_limit["value"]), 6) == round(20 * 290.2644471258237, 6)

    energy_rows = list(csv.DictReader((output_dir / "energy.csv").open()))
    assert any(
        row["metric_group"] == "site_energy" and row["end_use"] == "heating" and row["variant_id"] == "code_minimum"
        for row in energy_rows
    )
    assert any(
        row["metric_group"] == "phius_net_source_energy"
        and row["end_use"] == "total"
        and row["variant_id"] == "enerphit_by_demand"
        for row in energy_rows
    )

    demand_detail_rows = list(csv.DictReader((output_dir / "demand-detail.csv").open()))
    assert any(
        row["demand_type"] == "heating"
        and row["contribution_type"] == "loss"
        and row["item"] == "walls_ag"
        and row["variant_id"] == "code_minimum"
        for row in demand_detail_rows
    )
    assert any(
        row["demand_type"] == "cooling" and row["contribution_type"] == "limit" and row["variant_id"] == "code_minimum"
        for row in demand_detail_rows
    )


def test_scrape_linde_fixture_uses_dynamic_r_value_labels(scraped_linde: ScrapeOutput) -> None:
    output_dir = scraped_linde.data_dir

    assert "scraped PHPP 10.6: 5 variants, recommended=as_drawn" in scraped_linde.command_output

    rows = list(csv.DictReader((output_dir / "variants.csv").open()))
    r_value_rows = [row for row in rows if row["section"] == "r_values" and row["variant_id"] == "as_drawn"]
    r_value_ids = {row["field_id"] for row in r_value_rows}
    r_value_labels = {row["phpp_label"] for row in r_value_rows}

    assert "r_values.assembly_01" in r_value_ids
    assert "r_values.r_values" not in r_value_ids
    assert "r_values.assembly_02" not in r_value_ids
    assert "r_values.assembly_04" not in r_value_ids
    assert "01ud-c-W-CS - Crawlspace" in r_value_labels
    assert "06ud-g-R-FL - Flat" in r_value_labels
    assert "02ud-" not in r_value_labels

    envelope_rows = [row for row in rows if row["section"] == "envelope" and row["variant_id"] == "as_drawn"]
    envelope_ids = {row["field_id"] for row in envelope_rows}
    envelope_labels = {row["phpp_label"] for row in envelope_rows}

    assert "envelope.assembly_06" in envelope_ids
    assert "envelope.assembly_07" in envelope_ids
    assert "envelope.assembly_08" in envelope_ids
    assert "R-AT - Attic" in envelope_labels
    assert "R-FL - Flat" in envelope_labels
    assert "R-VT - Vaulted" in envelope_labels

    airflow_rows = list(csv.DictReader((output_dir / "room-airflows.csv").open()))
    assert len([row for row in airflow_rows if row["row_type"] == "room"]) == 28
    assert airflow_rows[-1]["room_name"] == "Totals"
    assert any(row["room_name"] == "113-BATH" for row in airflow_rows)
    assert any(row["room_name"] == "Kitchen Extract Hood - ON" for row in airflow_rows)

    building_metric_rows = list(csv.DictReader((output_dir / "building-metrics.csv").open()))
    assert len([row for row in building_metric_rows if row["metric"] == "treated_floor_area"]) == 5

    certification_rows = list(csv.DictReader((output_dir / "certification.csv").open()))
    assert any(row["metric"] == "per_demand" and row["variant_id"] == "as_drawn" for row in certification_rows)

    energy_rows = list(csv.DictReader((output_dir / "energy.csv").open()))
    assert any(row["metric_group"] == "per" and row["variant_id"] == "as_drawn" for row in energy_rows)

    demand_detail_rows = list(csv.DictReader((output_dir / "demand-detail.csv").open()))
    assert any(row["demand_type"] == "cooling" and row["variant_id"] == "as_drawn" for row in demand_detail_rows)


def test_scrape_unknown_version_fails_loudly(tmp_path: Path) -> None:
    workbook_path = tmp_path / "minimal.xlsx"
    _save_minimal_workbook(workbook_path)
    runner = CliRunner()

    result = runner.invoke(
        main,
        [
            "scrape",
            str(workbook_path),
            "--out",
            str(tmp_path / "data"),
            "--phpp-version",
            "10.7",
        ],
    )

    assert result.exit_code != 0
    assert "Unsupported PHPP version '10.7'" in result.output


def test_scrape_project_path_rejects_stale_project_schema_before_writing_data(tmp_path: Path) -> None:
    project = tmp_path / "Project" / "04_Web"
    project.mkdir(parents=True)
    (project / "project.yaml").write_text(yaml.safe_dump(_project_yaml("0.1.0"), sort_keys=False))
    runner = CliRunner()

    result = runner.invoke(main, ["scrape", str(project)])

    assert result.exit_code != 0
    assert "schema_version" in result.output
    assert "0.2.0" in result.output
    assert not (project / "data").exists()


def test_scrape_missing_workbook_fails_loudly(tmp_path: Path) -> None:
    runner = CliRunner()

    result = runner.invoke(main, ["scrape", str(tmp_path / "missing.xlsx"), "--out", str(tmp_path / "data")])

    assert result.exit_code != 0
    assert "does not exist" in result.output


def test_scrape_unreadable_workbook_fails_loudly(tmp_path: Path) -> None:
    workbook_path = tmp_path / "not-a-workbook.xlsx"
    workbook_path.write_text("not an xlsx file")
    runner = CliRunner()

    result = runner.invoke(main, ["scrape", str(workbook_path), "--out", str(tmp_path / "data")])

    assert result.exit_code != 0
    assert "File is not a zip file" in result.output


def test_scrape_missing_variants_sheet_fails_loudly(tmp_path: Path) -> None:
    workbook_path = tmp_path / "missing-variants.xlsx"
    _save_minimal_workbook(workbook_path)
    runner = CliRunner()

    result = runner.invoke(main, ["scrape", str(workbook_path), "--out", str(tmp_path / "data")])

    assert result.exit_code != 0
    assert "Worksheet Variants does not exist" in result.output


def test_scrape_empty_variant_set_fails_loudly(tmp_path: Path) -> None:
    workbook_path = tmp_path / "empty-variants.xlsx"
    workbook = _minimal_workbook()
    workbook.create_sheet("Variants")
    workbook.save(workbook_path)
    runner = CliRunner()

    result = runner.invoke(main, ["scrape", str(workbook_path), "--out", str(tmp_path / "data")])

    assert result.exit_code != 0
    assert "No active variants found" in result.output


def test_scrape_all_null_variant_data_fails_loudly(tmp_path: Path) -> None:
    workbook_path = tmp_path / "all-null-variant-data.xlsx"
    workbook = _minimal_workbook()
    variants = workbook.create_sheet("Variants")
    variants["E2"] = "1 - Empty Variant"
    workbook.save(workbook_path)
    runner = CliRunner()

    result = runner.invoke(main, ["scrape", str(workbook_path), "--out", str(tmp_path / "data")])

    assert result.exit_code != 0
    assert "No variant data found" in result.output


def test_variant_columns_follow_excel_column_order(tmp_path: Path) -> None:
    workbook_path = tmp_path / "variant-order.xlsx"
    workbook = _minimal_workbook()
    variants = workbook.create_sheet("Variants")
    variants["G2"] = "2 - Right"
    variants["E2"] = "1 - Left"
    variants["H2"] = "3 - Far Right"
    workbook.save(workbook_path)

    columns = OpenpyxlWorkbookReader(workbook_path).read_variant_columns(_minimal_schema())

    assert [column.name for column in columns] == ["Left", "Right", "Far Right"]
    assert [column.source_column for column in columns] == ["E", "G", "H"]


def _scrape_fixture(workbook_path: Path, output_dir: Path) -> ScrapeOutput:
    if not workbook_path.exists():
        pytest.skip(f"PHPP workbook fixture is not available: {workbook_path}")
    runner = CliRunner()
    result = runner.invoke(main, ["scrape", str(workbook_path), "--out", str(output_dir)])
    assert result.exit_code == 0, result.output
    return ScrapeOutput(data_dir=output_dir, command_output=result.output)


def _project_yaml(schema_version: str) -> dict[str, object]:
    return {
        "schema_version": schema_version,
        "slug": "project-2606",
        "project_title": "2606 Vandam",
        "client_name": "Client",
        "building_name": "Building",
        "phase": "Design Analysis",
        "report_date": "2026-05-21",
        "prepared_by": "BLDGTYP",
        "contact_email": "ed@bldgtyp.com",
        "target_standard": "TBD",
        "certification_program": "TBD",
        "certification_path": "TBD",
        "building": {
            "address": "TBD",
            "city": "TBD",
            "state": "TBD",
            "climate_zone": "TBD",
            "building_type": "TBD",
        },
        "source_files": {
            "phpp_path": "../07_PHPP/model.xlsx",
            "data_dir": "data",
            "assets_dir": "public/assets",
        },
        "publishing": {
            "production_url": "https://project-2606.bldgtyp.com",
            "cloudflare_pages_project": "bt-proj-2606-vandam",
        },
    }


def _assert_golden_csvs(actual_dir: Path, expected_dir: Path) -> None:
    for spec in REPORT_CSV_TABLES:
        assert (actual_dir / spec.filename).read_text().splitlines() == (
            expected_dir / spec.filename
        ).read_text().splitlines()


def _save_minimal_workbook(path: Path) -> None:
    _minimal_workbook().save(path)


def _minimal_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["B5"] = "10.6"
    return workbook


def _minimal_schema() -> WorkbookSchema:
    return WorkbookSchema(
        version="10.6",
        variant_sheet="Variants",
        climate_sheet="Climate",
        room_ventilation_sheet="Additional Ventilation",
        phpp_version_cell="B5",
        phpp_version_named_range="PHPP_Version",
        variant_header_row=2,
        variant_first_data_row=3,
        variants=(),
        climate_monthly=ClimateMonthlySchema(
            sheet="Climate",
            start_row=1,
            end_row=1,
            start_col="A",
            end_col="B",
        ),
        room_ventilation=RoomVentilationSchema(
            sheet="Additional Ventilation",
            header_col="A",
            header_label="",
            entry_col="A",
            first_entry_label="1",
            last_col="B",
        ),
    )
