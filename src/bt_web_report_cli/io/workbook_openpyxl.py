"""openpyxl-backed PHPP workbook reader."""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from bt_web_report_schemas.phpp.models import WorkbookSchema


@dataclass(frozen=True)
class VariantColumn:
    """One active variant column in the PHPP Variants worksheet."""

    id: str
    name: str
    order: int
    column_index: int
    source_column: str


class OpenpyxlWorkbookReader:
    """Read saved PHPP workbook values without launching Excel."""

    def __init__(self, path: Path) -> None:
        self.path = path
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="wmf image format is not supported.*")
            warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
            warnings.filterwarnings("ignore", message="Cannot parse header or footer.*")
            self._workbook = load_workbook(path, data_only=True, read_only=False)

    def detect_phpp_version(self) -> str:
        """Read PHPP version from named range, falling back to Data!B5."""

        value = self._read_named_range("PHPP_Version")
        if value in (None, ""):
            value = self._workbook["Data"]["B5"].value
        if value in (None, ""):
            msg = "Could not detect PHPP version from named range PHPP_Version or Data!B5."
            raise ValueError(msg)
        return str(value).strip()

    def read_variant_columns(self, schema: WorkbookSchema) -> tuple[VariantColumn, ...]:
        """Read active variant labels from the Variants header row."""

        sheet = self._workbook[schema.variant_sheet]
        columns: list[VariantColumn] = []
        for cell in sheet[schema.variant_header_row]:
            parsed = _parse_variant_header(cell.value)
            if parsed is None:
                continue
            order, name = parsed
            if order == 0:
                continue
            columns.append(
                VariantColumn(
                    id=_slugify(name),
                    name=name,
                    order=order,
                    column_index=cell.column,
                    source_column=get_column_letter(cell.column),
                )
            )

        if not columns:
            msg = f"No active variants found on {schema.variant_sheet} row {schema.variant_header_row}."
            raise ValueError(msg)
        return tuple(sorted(columns, key=lambda column: column.column_index))

    def read_variants(self, schema: WorkbookSchema, variant_columns: tuple[VariantColumn, ...]) -> list[dict[str, Any]]:
        """Read Variants worksheet rows into the first long-format report table."""

        sheet = self._workbook[schema.variant_sheet]
        rows: list[dict[str, Any]] = []
        for field in schema.variant_fields():
            if schema.section(field.section_id).start_row == field.row:
                continue
            if _clean_text(field.phpp_label) in {"", "-"}:
                continue
            datatype = sheet.cell(field.row, 3).value
            units = sheet.cell(field.row, 4).value
            phpp_label = _clean_text(datatype) if field.label_from_workbook else field.phpp_label.strip()
            if field.label_from_workbook and _is_placeholder_r_value(phpp_label):
                continue
            for variant in variant_columns:
                value = sheet.cell(field.row, variant.column_index).value
                if _is_blank_row(datatype, units, value):
                    continue
                rows.append(
                    {
                        "section": field.section_id,
                        "field_id": f"{field.section_id}.{field.id}",
                        "phpp_label": phpp_label,
                        "variant_id": variant.id,
                        "variant_name": variant.name,
                        "datatype": _clean_text(datatype) or field.phpp_label.strip(),
                        "units": _clean_text(units),
                        "value": value,
                        "excel_row": field.row,
                    }
                )
        return rows

    def read_climate_monthly(self, schema: WorkbookSchema) -> list[dict[str, Any]]:
        """Read active monthly Climate worksheet data in long format."""

        sheet = self._workbook[schema.climate_monthly.sheet]
        start_col = column_index_from_string(schema.climate_monthly.start_col)
        rows: list[dict[str, Any]] = []
        for row_number in range(schema.climate_monthly.start_row, schema.climate_monthly.end_row + 1):
            source_label = _clean_text(sheet.cell(row_number, start_col).value)
            mapping = _CLIMATE_ROW_MAP.get(source_label)
            if mapping is None:
                continue
            metric, orientation, units, factor, offset = mapping
            for month_index, month in enumerate(_MONTHS, start=1):
                value = sheet.cell(row_number, start_col + month_index).value
                if value in (None, ""):
                    continue
                rows.append(
                    {
                        "month": month,
                        "metric": metric,
                        "orientation": orientation,
                        "units": units,
                        "value": _convert_number(value, factor=factor, offset=offset),
                        "source_label": source_label,
                        "excel_row": row_number,
                    }
                )
        return rows

    def read_room_airflows(self, schema: WorkbookSchema) -> list[dict[str, Any]]:
        """Read Additional Ventilation room airflow rows plus totals."""

        room_schema = schema.room_ventilation
        sheet = self._workbook[room_schema.sheet]
        first_entry_row = self._find_room_first_entry_row(schema)
        last_entry_row = self._find_room_last_entry_row(schema, first_entry_row)

        rows: list[dict[str, Any]] = []
        for row_number in range(first_entry_row, last_entry_row + 1):
            room_name = _clean_text(sheet[f"E{row_number}"].value)
            if not room_name:
                continue
            rows.append(_room_airflow_row(sheet, row_number, room_name))

        if rows:
            rows.append(_room_airflow_totals(rows))
        return rows

    def _read_named_range(self, name: str) -> Any:
        defined_name = self._workbook.defined_names.get(name)
        if defined_name is None:
            return None
        for sheet_name, coord in defined_name.destinations:
            return self._workbook[sheet_name][coord].value
        return None

    def _find_room_first_entry_row(self, schema: WorkbookSchema) -> int:
        room_schema = schema.room_ventilation
        sheet = self._workbook[room_schema.sheet]
        header_row = None
        header_col = column_index_from_string(room_schema.header_col)
        entry_col = column_index_from_string(room_schema.entry_col)
        for row_number in range(1, 101):
            value = _clean_text(sheet.cell(row_number, header_col).value)
            if room_schema.header_label in value:
                header_row = row_number
                break
        if header_row is None:
            msg = f"Could not find Additional Ventilation room header '{room_schema.header_label}'."
            raise ValueError(msg)

        for row_number in range(header_row, header_row + 26):
            value = sheet.cell(row_number, entry_col).value
            try:
                entry_label = str(int(value))
            except (TypeError, ValueError):
                continue
            if entry_label == room_schema.first_entry_label:
                return row_number

        msg = f"Could not find first Additional Ventilation room entry after row {header_row}."
        raise ValueError(msg)

    def _find_room_last_entry_row(self, schema: WorkbookSchema, first_entry_row: int) -> int:
        room_schema = schema.room_ventilation
        sheet = self._workbook[room_schema.sheet]
        entry_col = column_index_from_string(room_schema.entry_col)
        for row_number in range(first_entry_row, first_entry_row + 500):
            if sheet.cell(row_number, entry_col).value is None:
                return row_number - 1
        msg = f"Could not find end of Additional Ventilation room entries after row {first_entry_row}."
        raise ValueError(msg)


def _parse_variant_header(value: object) -> tuple[int, str] | None:
    if not isinstance(value, str):
        return None
    match = re.match(r"\s*(\d+)\s*-\s*(.+?)\s*$", value)
    if match is None:
        return None
    name = match.group(2).strip()
    if not name:
        return None
    return int(match.group(1)), name


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "variant"


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(",", " ")


def _is_blank_row(datatype: object, units: object, value: object) -> bool:
    return _clean_text(datatype) in {"", "-"} and _clean_text(units) in {"", "-"} and value in (None, "", "-")


def _is_placeholder_r_value(label: str) -> bool:
    return re.fullmatch(r"\d{2}ud-", label) is not None


_MONTHS = (
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "oct",
    "nov",
    "dec",
)

_C_TO_F_OFFSET = 32.0
_C_TO_F_FACTOR = 9.0 / 5.0
_KWH_M2_TO_KWH_FT2 = 1.0 / 10.76391042

_CLIMATE_ROW_MAP = {
    "Exterior temperature": ("exterior_temperature", "", "degF", _C_TO_F_FACTOR, _C_TO_F_OFFSET),
    "Dew point temperature": ("dew_point_temperature", "", "degF", _C_TO_F_FACTOR, _C_TO_F_OFFSET),
    "Sky temperature": ("sky_temperature", "", "degF", _C_TO_F_FACTOR, _C_TO_F_OFFSET),
    "Radiation North": ("solar_radiation", "north", "kWh/ft2", _KWH_M2_TO_KWH_FT2, 0.0),
    "Radiation East": ("solar_radiation", "east", "kWh/ft2", _KWH_M2_TO_KWH_FT2, 0.0),
    "Radiation South": ("solar_radiation", "south", "kWh/ft2", _KWH_M2_TO_KWH_FT2, 0.0),
    "Radiation West": ("solar_radiation", "west", "kWh/ft2", _KWH_M2_TO_KWH_FT2, 0.0),
    "Horizontal radiation": ("solar_radiation", "horizontal", "kWh/ft2", _KWH_M2_TO_KWH_FT2, 0.0),
}

_M3H_TO_CFM = 0.588577779
_M3_TO_FT3 = 35.31466672
_M2_TO_FT2 = 10.76391042
_M_TO_FT = 3.280839895


def _convert_number(value: object, *, factor: float, offset: float = 0.0) -> object:
    if not isinstance(value, int | float):
        return value
    return value * factor + offset


def _room_airflow_row(sheet: Any, row_number: int, room_name: str) -> dict[str, Any]:
    v_supply = _number(sheet[f"J{row_number}"].value)
    v_extract = _number(sheet[f"K{row_number}"].value)
    reduction_high = _number(sheet[f"Q{row_number}"].value)
    reduction_med = _number(sheet[f"S{row_number}"].value)
    reduction_low = _number(sheet[f"U{row_number}"].value)

    return {
        "row_type": "room",
        "room_name": room_name,
        "amount": sheet[f"D{row_number}"].value,
        "allocation_to_vent_unit": sheet[f"F{row_number}"].value,
        "room_area_ft2": _number(sheet[f"G{row_number}"].value) * _M2_TO_FT2,
        "room_volume_ft3": _number(sheet[f"I{row_number}"].value) * _M3_TO_FT3,
        "room_height_ft": _number(sheet[f"H{row_number}"].value) * _M_TO_FT,
        "v_sup_high_cfm": v_supply * reduction_high * _M3H_TO_CFM,
        "v_eta_high_cfm": v_extract * reduction_high * _M3H_TO_CFM,
        "v_sup_med_cfm": v_supply * reduction_med * _M3H_TO_CFM,
        "v_eta_med_cfm": v_extract * reduction_med * _M3H_TO_CFM,
        "v_sup_low_cfm": v_supply * reduction_low * _M3H_TO_CFM,
        "v_eta_low_cfm": v_extract * reduction_low * _M3H_TO_CFM,
        "excel_row": row_number,
    }


def _room_airflow_totals(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total_fields = (
        "room_area_ft2",
        "room_volume_ft3",
        "v_sup_high_cfm",
        "v_eta_high_cfm",
        "v_sup_med_cfm",
        "v_eta_med_cfm",
        "v_sup_low_cfm",
        "v_eta_low_cfm",
    )
    totals: dict[str, Any] = {
        "row_type": "total",
        "room_name": "Totals",
        "amount": "",
        "allocation_to_vent_unit": "",
        "room_height_ft": "",
        "excel_row": "",
    }
    for field in total_fields:
        totals[field] = sum(_number(row[field]) for row in rows)
    return totals


def _number(value: object) -> float:
    if isinstance(value, int | float):
        return float(value)
    return 0.0
